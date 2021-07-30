# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import sys
from urllib.parse import urlparse, quote, urlencode, unquote
from urllib.request import urlopen
import re
 
def fetch(query_str):
    query = {'q': "".join(query_str)}   # list --> str: "".join(list)
    url = 'https://fanyi.youdao.com/openapi.do?keyfrom=11pegasus11&key=273646050&type=data&doctype=json&version=1.1&' + urlencode(query)
    response = urlopen(url, timeout=3)
    html = response.read().decode('utf-8')
    return html
 
def parse(html, num):
    d = json.loads(html)
    try:
        if d.get('errorCode') == 0:
            explains = d.get('basic').get('explains')
            result = str(explains).replace('\'', "").replace('[', "").replace(']', "")  #.replace真好用~
            sheet.cell(row=num, column=2).value = result
            num = num+1
            for i in explains:
                print(i)
        else:
            print('无法翻译!****')
            sheet.cell(row = num, column = 2).value = ' '       #若无法翻译，则空出来
            num = num + 1
    except:
        print('****翻译出错!')      #若无法翻译，则空出来
        sheet.cell(row = num, column = 2).value = ' '
        num = num + 1
 
def main():
    Sheet1 = ExcelFile['Sheet1']; num = 1
    while(1):
        word = Sheet1.cell(row = num+2, column = 1).value
        if(word != None):
            print('正在翻译第', end=''); print(num, end=''); print('个单词')
            print(word)
            parse(fetch(word), num)
            num += 1
            print()
        else:
            print('翻译结束！')
            break
    ExcelFile.close()
    out.save('out.xlsx')
 
if __name__ == '__main__':
    ExcelFile = load_workbook('words.xlsx')      #输入文件
    out = Workbook()
    sheet = out.active
    sheet.title = "out"
    main()